"""
smart_consolidator.py

Standalone data-cleaning + consolidation pipeline for messy exported
Excel/CSV reports (Trial Balance / Balance Sheet / P&L style exports --
the kind that has a report title, a company-name row, a date row, a
split/duplicated header, blank spacer columns, and a TOTAL line sitting
on top of the real data).

Turns any of that into ONE flat, analysis-ready table with this fixed
column layout, always:

    Company_Name | Account | Sub_Account | Category | Debit | Credit | Amount_Type | Notes

This file is completely standalone -- it does not import from, or get
imported by, any other file in DataFusion Platform. Nothing else in the
app is touched by adding this file.

-----------------------------------------------------------------------
USAGE
-----------------------------------------------------------------------
    from smart_consolidator import clean_file, consolidate_files

    # one messy file -> one clean DataFrame
    df = clean_file("GSR Foods II LLC.xlsx")

    # many messy files -> one clean, consolidated DataFrame
    df = consolidate_files(["GSR Foods II LLC.xlsx", "GSR Foods III LLC.xlsx"])
    df.to_excel("clean_output.xlsx", index=False)

Each function also accepts an explicit company_name= if you don't want
it auto-detected from the report title row.
-----------------------------------------------------------------------
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Optional, Sequence

import pandas as pd
from openpyxl import load_workbook


# --------------------------------------------------------------------
# 1. Header detection
# --------------------------------------------------------------------

# Words that mark a row as "this looks like a real column heading row",
# not a title/company/date/metadata row.
HEADER_KEYWORDS = [
    "debit", "credit", "amount", "balance", "date", "account",
    "particulars", "description", "narration", "voucher", "memo",
    "quantity", "qty", "rate", "code", "gstin", "party", "invoice",
]

# Row text that means "this whole row is a summary line, drop it" --
# whether it's the header row, or a stray total sitting mid-data.
TOTAL_ROW_RE = re.compile(r"^(grand\s+)?(sub[\s-]*)?total(s)?$", re.IGNORECASE)

# Duplicate-named columns from a dedup step (e.g. "Account", "Account (2)",
# "Account (3)") -- these are really ONE hierarchical field split apart.
DUP_SUFFIX_RE = re.compile(r"^(.*)\s\(\d+\)$")

CATEGORY_RULES = [
    (re.compile(r"\brent\b", re.I), "Expense"),
    (re.compile(r"\bexpense", re.I), "Expense"),
    (re.compile(r"\bloan\b", re.I), "Liability"),
    (re.compile(r"\bpayable\b", re.I), "Liability"),
    (re.compile(r"\bliabilit", re.I), "Liability"),
    (re.compile(r"\bcash\b", re.I), "Asset"),
    (re.compile(r"\bbank\b", re.I), "Asset"),
    (re.compile(r"\breceivable\b|\ba/r\b", re.I), "Asset"),
    (re.compile(r"\basset", re.I), "Asset"),
    (re.compile(r"\brevenue\b|\bsales\b|\bincome\b", re.I), "Income"),
    (re.compile(r"\bequity\b|\bcapital\b", re.I), "Equity"),
    (re.compile(r"\bpayroll\b|\bwages\b|\bsalar", re.I), "Expense"),
    (re.compile(r"\btax(es)?\b", re.I), "Tax"),
]


def _is_blank(v) -> bool:
    if v is None:
        return True
    try:
        if pd.isna(v):
            return True
    except (TypeError, ValueError):
        pass
    return str(v).strip() == ""


def _clean_text(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _read_raw_rows(path, sheet_name=None) -> list[list]:
    """Reads a sheet as a raw 2D list of values -- no header assumption,
    no dtype guessing. Works for .xlsx/.xlsm; falls back to pandas
    (covers .csv too)."""
    path = str(path)
    if path.lower().endswith((".xlsx", ".xlsm")):
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    # csv / xls / anything pandas can read
    df = pd.read_csv(path, header=None) if path.lower().endswith(".csv") \
        else pd.read_excel(path, header=None, sheet_name=sheet_name or 0)
    return df.values.tolist()


def _header_score(row: Sequence) -> int:
    score = 0
    for cell in row:
        text = _clean_text(cell).lower()
        if not text:
            continue
        if any(kw in text for kw in HEADER_KEYWORDS):
            score += 1
    return score


def _detect_header_row(rows: list[list], scan_limit: int = 20) -> int:
    """Returns the index of the row that looks most like the real
    column-heading row, searched only in the first `scan_limit` rows
    (title/company/date rows always sit near the top)."""
    best_idx, best_score = 0, -1
    for i, row in enumerate(rows[:scan_limit]):
        score = _header_score(row)
        if score > best_score:
            best_idx, best_score = i, score
    return best_idx


def _detect_company_name(rows: list[list], header_idx: int) -> Optional[str]:
    """The company name is almost always the first non-blank cell of the
    first non-blank row above the header (e.g. 'GSR Foods II, LLC')."""
    for row in rows[:header_idx]:
        for cell in row:
            if not _is_blank(cell):
                text = _clean_text(cell)
                # Skip rows that are clearly the report type / date line,
                # not a company name.
                if re.search(r"trial balance|balance sheet|profit|p&l|as of|report", text, re.I):
                    continue
                return text
    return None


# --------------------------------------------------------------------
# 2. Column setup: dedupe headers, drop blank columns, merge hierarchy
# --------------------------------------------------------------------

def _dedupe_headers(header_row: Sequence) -> list[str]:
    """Mirrors the app's own header-cleaning convention: blank/duplicate
    headers get a numbered suffix, e.g. two blank headers in a row
    become 'Column' and 'Column (2)'."""
    seen: dict[str, int] = {}
    out = []
    for i, cell in enumerate(header_row):
        name = _clean_text(cell) or f"Column{i + 1}"
        if name in seen:
            seen[name] += 1
            out.append(f"{name} ({seen[name]})")
        else:
            seen[name] = 1
            out.append(name)
    return out


def _drop_fully_blank_columns(df: pd.DataFrame) -> pd.DataFrame:
    keep = [c for c in df.columns if not (df[c].map(_is_blank)).all()]
    return df[keep]


def _drop_fully_blank_rows(df: pd.DataFrame) -> pd.DataFrame:
    mask = ~df.apply(lambda r: all(_is_blank(v) for v in r), axis=1)
    return df[mask]


def _merge_hierarchical_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Finds groups like Account / Account (2) / Account (3) -- columns
    that only exist because of header-dedup, and are really one
    hierarchical field spread across several columns (common in Trial
    Balance exports: broad category in the first, specific ledger
    account in the last). Folds each group into:
        <name>       -> the FIRST non-blank value per row (broadest level)
        Sub_Account  -> the LAST non-blank value per row (most specific)
    Only the "Account"-flavoured group is mapped to Sub_Account by name;
    any other duplicate-suffix group is folded the same way but keeps
    its own base name (first-non-blank wins, no data lost).
    """
    cols = list(df.columns)
    groups: dict[str, list[str]] = {}
    order: list[str] = []
    for c in cols:
        m = DUP_SUFFIX_RE.match(str(c))
        base = m.group(1) if m else str(c)
        if base not in groups:
            groups[base] = []
            order.append(base)
        groups[base].append(c)

    out = pd.DataFrame(index=df.index)
    sub_account_col = None
    for base in order:
        members = groups[base]
        if len(members) == 1:
            out[base] = df[members[0]]
            continue
        sub_df = df[members]
        first_vals = sub_df.apply(
            lambda r: next((v for v in r if not _is_blank(v)), ""), axis=1
        )
        last_vals = sub_df.apply(
            lambda r: next((v for v in reversed(list(r)) if not _is_blank(v)), ""), axis=1
        )
        out[base] = first_vals
        if re.search(r"account", base, re.I):
            sub_account_col = last_vals
        else:
            # Not clearly an "Account" hierarchy -- still avoid losing
            # data: if the last value differs from the first, keep it
            # in a "<base> Detail" column instead of silently dropping it.
            differs = (last_vals.astype(str) != first_vals.astype(str)) & (last_vals != "")
            if differs.any():
                out[f"{base} Detail"] = last_vals.where(differs, "")

    if sub_account_col is not None:
        main_vals = out.get("Account", pd.Series("", index=df.index))
        sub_account_col = sub_account_col.where(sub_account_col.astype(str) != main_vals.astype(str), "")
        out["Sub_Account"] = sub_account_col
    return out


# --------------------------------------------------------------------
# 3. Debit/Credit normalization + category detection
# --------------------------------------------------------------------

def _to_number(v) -> float:
    if _is_blank(v):
        return 0.0
    text = str(v).strip()
    neg = text.startswith("(") and text.endswith(")")
    text = re.sub(r"[^\d.\-]", "", text.replace(",", ""))
    if text in ("", "-", "."):
        return 0.0
    try:
        num = float(text)
    except ValueError:
        return 0.0
    return -num if neg else num


def _find_col(columns: Sequence[str], *keywords: str) -> Optional[str]:
    for c in columns:
        low = str(c).lower()
        if any(kw in low for kw in keywords):
            return c
    return None


def _detect_category(text: str) -> str:
    for pattern, category in CATEGORY_RULES:
        if pattern.search(text):
            return category
    return "Other"


# --------------------------------------------------------------------
# 4. Main pipeline
# --------------------------------------------------------------------

def clean_and_consolidate(rows: list[list], company_name: Optional[str] = None) -> pd.DataFrame:
    """Runs the full pipeline on one sheet's raw rows and returns the
    final flat table:
        Company_Name | Account | Sub_Account | Category | Debit | Credit | Amount_Type | Notes
    """
    if not rows:
        return pd.DataFrame(columns=[
            "Company_Name", "Account", "Sub_Account", "Category",
            "Debit", "Credit", "Amount_Type", "Notes",
        ])

    # ---- 1. detect header row, remove everything above it (metadata) ----
    header_idx = _detect_header_row(rows)
    detected_company = company_name or _detect_company_name(rows, header_idx)
    header = _dedupe_headers(rows[header_idx])
    data_rows = rows[header_idx + 1:]

    df = pd.DataFrame(data_rows, columns=header)

    # ---- 2. remove blank rows/cols, duplicate header rows, TOTAL rows ----
    df = _drop_fully_blank_rows(df)
    header_set = {_clean_text(h).lower() for h in header}
    is_repeated_header = df.apply(
        lambda r: {_clean_text(v).lower() for v in r if not _is_blank(v)} <= header_set
        and any(not _is_blank(v) for v in r),
        axis=1,
    )
    df = df[~is_repeated_header]
    is_total_row = df.apply(
        lambda r: any(TOTAL_ROW_RE.match(_clean_text(v)) for v in r), axis=1
    )
    df = df[~is_total_row]
    df = _drop_fully_blank_columns(df)

    # ---- 3. trim/standardize every text cell ----
    for c in df.columns:
        df[c] = df[c].map(lambda v: v if isinstance(v, (int, float)) else _clean_text(v))

    # ---- 4. merge hierarchical Account / Account (2) / Account (3) ... ----
    df = _merge_hierarchical_columns(df)

    # ---- 5. locate Debit/Credit, build Amount_Type ----
    debit_col = _find_col(df.columns, "debit")
    credit_col = _find_col(df.columns, "credit")
    debit = df[debit_col].map(_to_number) if debit_col else pd.Series(0.0, index=df.index)
    credit = df[credit_col].map(_to_number) if credit_col else pd.Series(0.0, index=df.index)
    amount_type = [
        "Debit" if d > 0 else ("Credit" if c > 0 else "")
        for d, c in zip(debit, credit)
    ]

    # ---- 6. Account / Sub_Account / Category ----
    account_col = _find_col(
        [c for c in df.columns if c != "Sub_Account"], "account", "particulars", "description"
    )
    account = df[account_col] if account_col else pd.Series("", index=df.index)
    sub_account = df["Sub_Account"] if "Sub_Account" in df.columns else pd.Series("", index=df.index)
    category = [
        _detect_category(f"{a} {s}") for a, s in zip(account, sub_account)
    ]

    # ---- 7. anything left over (memo/voucher/etc.) goes in Notes ----
    used = {account_col, "Sub_Account", debit_col, credit_col}
    used.discard(None)
    leftover_cols = [c for c in df.columns if c not in used]
    if leftover_cols:
        notes = df[leftover_cols].apply(
            lambda r: " | ".join(str(v) for v in r if not _is_blank(v)), axis=1
        )
    else:
        notes = pd.Series("", index=df.index)

    out = pd.DataFrame({
        "Company_Name": detected_company or "",
        "Account": account.values,
        "Sub_Account": sub_account.values,
        "Category": category,
        "Debit": debit.values,
        "Credit": credit.values,
        "Amount_Type": amount_type,
        "Notes": notes.values,
    })

    # ---- 8. final safety pass: drop dead rows/dupes ----
    out = out[~((out["Account"] == "") & (out["Sub_Account"] == "") & (out["Debit"] == 0) & (out["Credit"] == 0))]
    out = out.drop_duplicates().reset_index(drop=True)
    return out


def clean_file(path, sheet_name=None, company_name: Optional[str] = None) -> pd.DataFrame:
    """Clean a single Excel/CSV file (optionally a specific sheet) into
    the flat output table."""
    rows = _read_raw_rows(path, sheet_name)
    return clean_and_consolidate(rows, company_name=company_name)


def consolidate_files(paths: Sequence, company_names: Optional[Sequence[Optional[str]]] = None) -> pd.DataFrame:
    """Clean and merge several messy files into one consolidated table."""
    names = company_names or [None] * len(paths)
    frames = [clean_file(p, company_name=n) for p, n in zip(paths, names)]
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()