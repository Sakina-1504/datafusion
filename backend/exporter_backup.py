"""
exporter.py

Turns the app's in-memory data (consolidated rows, GST summary, party
summary, validation issues, filtered views) into a single, properly
formatted .xlsx workbook -- the kind a CA would actually want to open:
bold coloured headers, frozen header row, autofilter, sensible column
widths, number formatting on amount columns, and a bold Grand Total
row at the bottom of every data sheet.

Also builds a "Source Files" reference sheet (every sub file/company
that fed the report), a Credit = Debit checkpoint on the Summary
sheet, and turns each "Source File" cell on Consolidated Data into a
clickable link back to its row on the Source Files sheet.

Also provides open_file()/open_containing_folder() helpers so the UI
can take the user straight into Excel after export, cross-platform.
"""

import os
import re
import platform
import subprocess
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.filters import numeric_columns, grand_totals
from backend.column_mapper import detect_columns
from backend.settings import _user_data_dir

# Text shown whenever a column has no real heading in the source file --
# kept as one constant so it's identical everywhere it's used.
NO_HEADING_TEXT = "No heading in source file"
LINK_FONT = Font(color="1155CC", underline="single", size=11)

HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16, color="1E3A8A")
HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"),
)


def _to_number(value, default=0.0):
    try:
        if value == "" or value is None:
            return default
        return float(value)
    except (ValueError, TypeError):
        return default


def _credit_debit_checkpoint(rows, columns):
    """Checks whether Total Debit equals Total Credit across every
    consolidated row -- a common tie-out check before a report is
    considered final. Returns None if the source data doesn't have
    both a debit and a credit column, so the section is simply
    skipped rather than showing a false mismatch."""
    mapping = detect_columns(columns)
    debit_col = mapping.get("debit")
    credit_col = mapping.get("credit")
    if not debit_col or not credit_col:
        return None
    total_debit = sum(_to_number(row.get(debit_col)) for row in rows)
    total_credit = sum(_to_number(row.get(credit_col)) for row in rows)
    difference = round(total_debit - total_credit, 2)
    return {
        "debit_col": debit_col,
        "credit_col": credit_col,
        "total_debit": round(total_debit, 2),
        "total_credit": round(total_credit, 2),
        "difference": difference,
        "matched": abs(difference) < 0.01,
    }


def _source_files_breakdown(rows):
    """Groups consolidated rows by the file they came from, so each
    'sub file' that went into the report can be listed on its own
    reference sheet -- file name, company name assigned to it, and
    how many rows it contributed."""
    breakdown = {}
    order = []
    for row in rows:
        fname = row.get("__source_file", "Unknown")
        if fname not in breakdown:
            breakdown[fname] = {"company_name": row.get("Company Name", ""), "row_count": 0}
            order.append(fname)
        breakdown[fname]["row_count"] += 1
    return order, breakdown


def _autofit_columns(ws, columns, rows, min_width=10, max_width=45):
    for idx, col in enumerate(columns, start=1):
        longest = len(str(col))
        for row in rows[:500]:
            val_len = len(str(row.get(col, "")))
            if val_len > longest:
                longest = val_len
        ws.column_dimensions[get_column_letter(idx)].width = max(min_width, min(longest + 3, max_width))


GROUP_SHADE_FILLS = [
    PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
    PatternFill(start_color="EAF1FB", end_color="EAF1FB", fill_type="solid"),
]
BOUNDARY_TOP_SIDE = Side(style="medium", color="1E3A8A")


def _shade_source_file_groups(ws, rows, num_columns, start_row=1):
    """Purely cosmetic pass over the Consolidated Data sheet: doesn't
    touch any cell value, doesn't add/remove rows. It just alternates
    a light background tint each time the source file changes, and
    draws a bold top border on the first row of every new file's
    block -- so it's visually obvious where one source file's data
    ends and the next one's begins."""
    prev_file = None
    group_idx = -1
    r = start_row
    for row in rows:
        r += 1
        fname = row.get("__source_file", "")
        is_boundary = prev_file is not None and fname != prev_file
        if fname != prev_file:
            group_idx += 1
            prev_file = fname
        fill = GROUP_SHADE_FILLS[group_idx % 2]
        for c_idx in range(1, num_columns + 1):
            cell = ws.cell(row=r, column=c_idx)
            cell.fill = fill
            if is_boundary:
                existing = cell.border
                cell.border = Border(
                    left=existing.left, right=existing.right,
                    top=BOUNDARY_TOP_SIDE, bottom=existing.bottom,
                )


def _highlight_total_rows(ws, columns, rows, start_row=1):
    """A row whose Account cell literally reads "TOTAL" is a
    per-company subtotal carried over from the original source file
    (e.g. QuickBooks trial balances often have one). Left as-is it
    sits squeezed into one narrow column while the column right next
    to it is empty for that row. This merges the "TOTAL" label into
    that empty space beside it (so it reads across both columns
    instead of looking cramped) and draws a line directly above and
    below the row -- it does NOT touch the row's fill color, which
    stays whatever the source-file group shading already set it to."""
    r = start_row
    for row in rows:
        r += 1
        total_idx = None
        for idx, col in enumerate(columns):
            if str(row.get(col, "")).strip().upper() == "TOTAL":
                total_idx = idx
                break
        if total_idx is None:
            continue

        # Shift the label into the next column if it's empty there.
        if total_idx + 1 < len(columns):
            next_col = columns[total_idx + 1]
            if str(row.get(next_col, "")).strip() == "":
                first_c = total_idx + 1   # 1-based openpyxl column of the TOTAL cell
                second_c = total_idx + 2  # 1-based openpyxl column of the empty cell beside it
                ws.merge_cells(start_row=r, start_column=first_c, end_row=r, end_column=second_c)
                merged_cell = ws.cell(row=r, column=first_c)
                merged_cell.value = "TOTAL"
                merged_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Line above and below the row -- no fill/font change at all.
        for c_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c_idx)
            existing = cell.border
            cell.border = Border(
                left=existing.left, right=existing.right,
                top=BOUNDARY_TOP_SIDE, bottom=BOUNDARY_TOP_SIDE,
            )


def _write_table(ws, columns, rows, start_row=1, add_grand_total=True):
    for c_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=start_row, column=c_idx, value=str(col))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    num_cols = numeric_columns(rows, columns, sample_size=len(rows) or 1)

    r = start_row
    for row in rows:
        r += 1
        for c_idx, col in enumerate(columns, start=1):
            value = row.get(col, "")
            cell = ws.cell(row=r, column=c_idx, value=value)
            cell.border = THIN_BORDER
            if col in num_cols and value not in ("", None):
                try:
                    cell.value = float(value)
                    cell.number_format = "#,##0.00"
                except (ValueError, TypeError):
                    pass

    if add_grand_total and rows:
        r += 1
        totals = grand_totals(rows, columns)
        label_written = False
        for c_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=c_idx)
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            cell.border = THIN_BORDER
            if col in totals:
                cell.value = totals[col]
                cell.number_format = "#,##0.00"
            elif not label_written:
                cell.value = "GRAND TOTAL"
                label_written = True

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    if rows:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(columns))}{start_row + len(rows)}"
    _autofit_columns(ws, columns, rows)
    return r + 2


def export_full_report(output_path, consolidated, gst_summary=None, party_summary=None,
                        validation_issues=None, source_files=None, company_summary=None,
                        month_summary=None, consolidated_at=None, sections=None, file_paths=None):
    """
    Builds the complete multi-sheet workbook:
      1. Index               - what's in this workbook and why (so nobody
                                has to guess what each tab is for)
      2. Summary              - headline numbers, which files went in, and
                                the Credit = Debit checkpoint
      3. Source Files         - every sub file merged in, with its Company
                                Name and row count (linked from Consolidated Data)
      4. Consolidated Data    - every merged row + Grand Total row, with a
                                clickable Source File reference
      5. Company Summary      - totals per company (from the always-filled
                                 "Company Name" column)
      6. GST Summary          - rate-wise, CGST/SGST/IGST totals
      7. Monthly Summary      - totals per month, for filing/MIS trend
      8. Party Summary        - customer/vendor wise totals
      9. Issues Found         - every validation problem, with reference
    Any section is skipped gracefully if its data wasn't supplied.
    """

    sections = sections or {}
    file_paths = file_paths or {}

    wb = Workbook()
    rows = consolidated.get("rows", [])
    columns = [c for c in consolidated.get("columns", []) if not c.startswith("__")]
    # Guard against any stray blank header slipping through to export --
    # every column must show something a user can understand.
    columns = [c if str(c).strip() else NO_HEADING_TEXT for c in columns]
    # Safety net: a column with no real heading is never shown in the
    # export, full stop -- even if it somehow reached this point still
    # carrying data (e.g. from an already-imported session predating an
    # import-side fix). Drop the column and strip it out of every row.
    no_heading_cols = [c for c in columns if "no heading" in str(c).lower()]
    if no_heading_cols:
        columns = [c for c in columns if c not in no_heading_cols]
        rows = [{k: v for k, v in row.items() if k not in no_heading_cols} for row in rows]
    # Safety net #2: a column can look fine in any one file on its own
    # (so the per-file import cleanup leaves it alone) but still turn
    # out completely blank once every file is merged together -- e.g.
    # a column named "Account" that no source file in this batch ever
    # actually put data into. Check across the WHOLE merged dataset and
    # drop any column that's empty in every single row, from every file.
    if rows:
        fully_empty_cols = [
            c for c in columns
            if all(str(row.get(c, "")).strip() == "" for row in rows)
        ]
        if fully_empty_cols:
            columns = [c for c in columns if c not in fully_empty_cols]
            rows = [{k: v for k, v in row.items() if k not in fully_empty_cols} for row in rows]
    # Company Name always leads the sheet since every row is tagged with it.
    if "Company Name" in columns:
        columns = ["Company Name"] + [c for c in columns if c != "Company Name"]
    has_source = bool(rows) and "__source_file" in rows[0]
    display_columns = columns + (["Source File"] if has_source else [])

    consolidated_at_text = (consolidated_at or datetime.now()).strftime("%d-%b-%Y %I:%M %p")

    # ---------- 1. Index / how-to-use sheet ----------
    ws_idx = wb.active
    ws_idx.title = "Index"
    ws_idx["A1"] = "DataFusion Platform -- Report Guide"
    ws_idx["A1"].font = TITLE_FONT
    ws_idx["A2"] = f"Consolidated on: {consolidated_at_text}   |   Exported on: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"
    ws_idx["A2"].font = Font(italic=True, size=10, color="666666")

    guide_rows = [
        ("Summary", "Headline totals, files used, credit/debit checkpoint, CGST/SGST/IGST and data-quality snapshot."),
        ("Source Files", "Every source file (sub file) that went into this report, with the Company Name assigned to it and how many rows it contributed. Click a 'Source File' cell on Consolidated Data to jump here."),
        ("Consolidated Data", "Every row from every imported file, merged, with a Company Name column, a clickable Source File reference, and a Grand Total row."),
        ("Company Summary", "Taxable value, tax and row count per company -- useful when you're handling more than one client at once."),
        ("GST Summary", "Rate-wise (5%/12%/18%/28%) taxable and tax totals, ready for GSTR-1 rate-wise filing."),
        ("Monthly Summary", "Taxable and tax totals grouped by month -- for filing trend or a management report."),
        ("Party Summary", "Taxable value, tax and invoice count per customer/vendor."),
        ("Issues Found", "Every data-quality issue detected (missing fields, invalid GSTIN, duplicates, tax mismatches) with its exact file/sheet/row reference."),
    ]
    r = 4
    header_row = r
    for c_idx, text in ((1, "Sheet"), (2, "What it's for")):
        cell = ws_idx.cell(row=r, column=c_idx, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for name, desc in guide_rows:
        r += 1
        band_fill = GROUP_SHADE_FILLS[(r - header_row) % 2]
        name_cell = ws_idx.cell(row=r, column=1, value=name)
        name_cell.font = Font(bold=True)
        name_cell.border = THIN_BORDER
        name_cell.fill = band_fill
        name_cell.alignment = Alignment(vertical="center")
        desc_cell = ws_idx.cell(row=r, column=2, value=desc)
        desc_cell.border = THIN_BORDER
        desc_cell.fill = band_fill
        desc_cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws_idx.column_dimensions["A"].width = 22
    ws_idx.column_dimensions["B"].width = 95
    ws_idx.freeze_panes = ws_idx.cell(row=header_row + 1, column=1)

    # ---------- 2. Summary sheet ----------
    ws = wb.create_sheet("Summary")
    ws["A1"] = "DataFusion Platform -- Consolidation Summary"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Consolidated on: {consolidated_at_text}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws["A3"] = f"Report generated/exported on: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"
    ws["A3"].font = Font(italic=True, size=10, color="666666")

    r = 5
    ws.cell(row=r, column=1, value="Files included in this export:").font = Font(bold=True)
    r += 1
    for f in (source_files or []):
        ws.cell(row=r, column=1, value=f"  - {f}")
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Grand Totals").font = Font(bold=True, size=13, color="1E3A8A")
    r += 1
    totals = grand_totals(rows, columns)
    ws.cell(row=r, column=1, value="Total Rows").font = Font(bold=True)
    ws.cell(row=r, column=2, value=len(rows))
    r += 1
    for col, total in totals.items():
        ws.cell(row=r, column=1, value=f"Total {col}").font = Font(bold=True)
        cell = ws.cell(row=r, column=2, value=total)
        cell.number_format = "#,##0.00"
        r += 1

    checkpoint = _credit_debit_checkpoint(rows, columns)
    if checkpoint is not None and sections.get("checkpoint", True):
        r += 1
        ws.cell(row=r, column=1, value="Checkpoint: Credit = Debit").font = Font(bold=True, size=13, color="1E3A8A")
        r += 1
        ws.cell(row=r, column=1, value=f"Total {checkpoint['debit_col']}").font = Font(bold=True)
        ws.cell(row=r, column=2, value=checkpoint["total_debit"]).number_format = "#,##0.00"
        r += 1
        ws.cell(row=r, column=1, value=f"Total {checkpoint['credit_col']}").font = Font(bold=True)
        ws.cell(row=r, column=2, value=checkpoint["total_credit"]).number_format = "#,##0.00"
        r += 1
        ws.cell(row=r, column=1, value="Difference (Debit - Credit)").font = Font(bold=True)
        ws.cell(row=r, column=2, value=checkpoint["difference"]).number_format = "#,##0.00"
        r += 1
        colr = "16A34A" if checkpoint["matched"] else "B91C1C"
        status_text = "\u2714 Matched -- Credit and Debit tie out" if checkpoint["matched"] else "\u26A0 Mismatch -- Credit and Debit do not tie out"
        ws.cell(row=r, column=1, value="Checkpoint Status").font = Font(bold=True, color=colr)
        ws.cell(row=r, column=2, value=status_text).font = Font(bold=True, color=colr)

    if gst_summary and "error" not in gst_summary and sections.get("gst", True):
        r += 1
        ws.cell(row=r, column=1, value="CGST Total").font = Font(bold=True)
        ws.cell(row=r, column=2, value=gst_summary.get("cgst_total", 0)).number_format = "#,##0.00"
        r += 1
        ws.cell(row=r, column=1, value="SGST Total").font = Font(bold=True)
        ws.cell(row=r, column=2, value=gst_summary.get("sgst_total", 0)).number_format = "#,##0.00"
        r += 1
        ws.cell(row=r, column=1, value="IGST Total").font = Font(bold=True)
        ws.cell(row=r, column=2, value=gst_summary.get("igst_total", 0)).number_format = "#,##0.00"
        r += 1
        mismatches = gst_summary.get("tax_mismatches", [])
        colr = "B91C1C" if mismatches else "16A34A"
        ws.cell(row=r, column=1, value="Tax Mismatches Found").font = Font(bold=True, color=colr)
        ws.cell(row=r, column=2, value=len(mismatches))

    if validation_issues is not None:
        r += 2
        if isinstance(validation_issues, dict):
            total_issues = sum(len(v) for v in validation_issues.values())
        else:
            total_issues = len(validation_issues)
        colr = "B91C1C" if total_issues else "16A34A"
        ws.cell(row=r, column=1, value="Data Quality Issues Found").font = Font(bold=True, color=colr)
        ws.cell(row=r, column=2, value=total_issues)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20

    # ---------- 3. Source Files sheet (sub files that fed this report) ----------
    file_order, file_breakdown = _source_files_breakdown(rows)
    file_row_map = {}
    if file_order and sections.get("source_files", True):
        ws_src = wb.create_sheet("Source Files")
        ws_src["A1"] = "Source Files -- every sub file merged into this report"
        ws_src["A1"].font = TITLE_FONT
        src_columns = ["Sr No", "Source File", "Company Name", "Row Count"]
        for c_idx, col in enumerate(src_columns, start=1):
            cell = ws_src.cell(row=3, column=c_idx, value=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        r_src = 3
        for i, fname in enumerate(file_order, start=1):
            r_src += 1
            info = file_breakdown[fname]
            ws_src.cell(row=r_src, column=1, value=i).border = THIN_BORDER
            name_cell = ws_src.cell(row=r_src, column=2, value=fname)
            name_cell.border = THIN_BORDER
            src_path = file_paths.get(fname)
            if src_path:
                # Clicking the filename opens the actual source workbook on disk.
                name_cell.hyperlink = f"file:///{src_path}"
                name_cell.font = LINK_FONT
            ws_src.cell(row=r_src, column=3, value=info["company_name"]).border = THIN_BORDER
            ws_src.cell(row=r_src, column=4, value=info["row_count"]).border = THIN_BORDER
            file_row_map[fname] = r_src
        ws_src.freeze_panes = "A4"
        ws_src.column_dimensions["A"].width = 8
        ws_src.column_dimensions["B"].width = 45
        ws_src.column_dimensions["C"].width = 30
        ws_src.column_dimensions["D"].width = 14

    # ---------- 4. Consolidated Data sheet ----------
    ws2 = wb.create_sheet("Consolidated Data")
    export_rows = []
    for row in rows:
        flat = {c: row.get(c, "") for c in columns}
        if "Source File" in display_columns:
            flat["Source File"] = row.get("__source_file", "")
        export_rows.append(flat)
    last_data_row = _write_table(ws2, display_columns, export_rows)

    # Cosmetic only -- shade rows by source file and mark where a new
    # file's data starts, so different files are easy to tell apart.
    _shade_source_file_groups(ws2, rows, len(display_columns))

    # Cosmetic only -- per-company "TOTAL" rows get their label merged
    # into the empty cell beside it and the whole row highlighted.
    _highlight_total_rows(ws2, display_columns, rows)

    # Turn each "Source File" cell into a clickable reference back to
    # that file's row on the Source Files sheet.
    if "Source File" in display_columns and file_row_map:
        src_col_idx = display_columns.index("Source File") + 1
        for r_idx, row in enumerate(export_rows, start=2):
            fname = row.get("Source File", "")
            target_row = file_row_map.get(fname)
            if target_row:
                cell = ws2.cell(row=r_idx, column=src_col_idx)
                cell.hyperlink = f"#'Source Files'!B{target_row}"
                cell.font = LINK_FONT

    # ---------- 5. Company Summary sheet ----------
    if (company_summary and "error" not in company_summary and company_summary.get("companies")
            and sections.get("company", True)):
        ws_co = wb.create_sheet("Company Summary")
        co_rows = []
        for name, vals in company_summary["companies"].items():
            co_rows.append({
                "Company Name": name,
                "Taxable Total": vals["taxable_total"],
                "Tax Total": vals["tax_total"],
                "Row Count": vals["row_count"],
            })
        _write_table(ws_co, ["Company Name", "Taxable Total", "Tax Total", "Row Count"], co_rows, add_grand_total=False)

    # ---------- 6. GST Summary sheet ----------
    if gst_summary and "error" not in gst_summary and sections.get("gst", True):
        ws3 = wb.create_sheet("GST Summary")
        gst_rows = []
        for rate, vals in sorted(gst_summary.get("rate_wise", {}).items()):
            gst_rows.append({
                "GST Rate (%)": rate,
                "Taxable Total": vals["taxable_total"],
                "Tax Total": vals["tax_total"],
                "Row Count": vals["row_count"],
            })
        _write_table(ws3, ["GST Rate (%)", "Taxable Total", "Tax Total", "Row Count"], gst_rows, add_grand_total=False)

    # ---------- 7. Monthly Summary sheet ----------
    if (month_summary and "error" not in month_summary and month_summary.get("months")
            and sections.get("monthly", False)):
        ws_mo = wb.create_sheet("Monthly Summary")
        month_rows = []
        for month, vals in sorted(month_summary["months"].items()):
            month_rows.append({
                "Month": month,
                "Taxable Total": vals["taxable_total"],
                "Tax Total": vals["tax_total"],
                "Row Count": vals["row_count"],
            })
        _write_table(ws_mo, ["Month", "Taxable Total", "Tax Total", "Row Count"], month_rows, add_grand_total=False)

    # ---------- 8. Party Summary sheet ----------
    if party_summary and "error" not in party_summary and sections.get("party", True):
        ws4 = wb.create_sheet("Party Summary")
        party_rows = []
        for name, vals in party_summary.get("parties", {}).items():
            party_rows.append({
                "Party Name": name,
                "Taxable Total": vals["taxable_total"],
                "Tax Total": vals["tax_total"],
                "Invoice Count": vals["invoice_count"],
            })
        _write_table(ws4, ["Party Name", "Taxable Total", "Tax Total", "Invoice Count"], party_rows, add_grand_total=False)

    # ---------- 9. Issues Found sheet ----------
    if validation_issues:
        ws5 = wb.create_sheet("Issues Found")
        issue_rows = []
        for issue_type, items in validation_issues.items():
            for item in items:
                issue_rows.append({
                    "Issue Type": issue_type.replace("_", " ").title(),
                    "Details": str(item),
                })
        if issue_rows:
            _write_table(ws5, ["Issue Type", "Details"], issue_rows, add_grand_total=False)
        else:
            ws5.cell(row=1, column=1, value="No issues found").font = Font(bold=True, color="16A34A")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


def export_filtered_view(output_path, columns, rows, title="Filtered Data"):
    """Simpler single-sheet export used by the Filters screen and the
    Analytics screen (exporting whatever the user is currently looking at)."""
    wb = Workbook()
    ws = wb.active
    ws.title = (title[:31] if title else "Filtered Data")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated on: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}  |  {len(rows)} row(s)"
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    display_columns = [c if str(c).strip() else NO_HEADING_TEXT for c in columns if not str(c).startswith("__")]
    _write_table(ws, display_columns, rows, start_row=4)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


def export_pivot_view(output_path, pivot_columns, pivot_rows, title="Pivot Table"):
    """Exports a pivot-shaped table (list of flat dict rows) with the
    same styling as everything else."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pivot Table"
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated on: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    _write_table(ws, pivot_columns, pivot_rows, start_row=4, add_grand_total=False)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


# ---------------- Cross-platform "open file / folder" helpers ---------------- #

def open_file(path):
    """Opens the given file with whatever application the OS has
    associated with .xlsx (i.e. Excel, if installed) -- this is what
    takes the user directly into Excel after export."""
    try:
        system = platform.system()
        if system == "Windows":
            os.startfile(path)  # noqa: only exists on Windows
        elif system == "Darwin":
            subprocess.run(["open", path], check=False)
        else:
            subprocess.run(["xdg-open", path], check=False)
        return True
    except Exception:
        return False


def open_containing_folder(path):
    """Opens the folder containing the file (selecting the file itself
    where the OS supports it)."""
    folder = os.path.dirname(os.path.abspath(path))
    try:
        system = platform.system()
        if system == "Windows":
            subprocess.run(["explorer", "/select,", os.path.abspath(path)], check=False)
        elif system == "Darwin":
            subprocess.run(["open", "-R", os.path.abspath(path)], check=False)
        else:
            subprocess.run(["xdg-open", folder], check=False)
        return True
    except Exception:
        return False


def open_file_with_highlighted_columns(file_path, sheet_name, columns_to_highlight):
    """Used by the Debit/Credit Checkpoint screen: makes a copy of the
    given source file with the named column(s) highlighted in yellow
    on the given sheet, then opens that copy in Excel -- so a mismatch
    can be traced straight back to the exact columns it came from.

    A copy is opened rather than the original file so the user's
    source file on disk is never modified. Copies are written to
    <app data>/highlighted_references and reused/overwritten on
    repeat opens of the same file/sheet/columns combination.

    Returns True if the highlighted copy was created and opened,
    False if the source file, sheet, or columns could not be found.
    """
    if not file_path or not os.path.exists(file_path):
        return False

    wanted = {str(c).strip() for c in (columns_to_highlight or []) if c}
    if not wanted:
        return False

    try:
        wb = load_workbook(file_path)
    except Exception:
        return False

    if sheet_name not in wb.sheetnames:
        return False
    ws = wb[sheet_name]

    header_row = next(ws.iter_rows(min_row=1, max_row=1), ())
    col_indexes = [cell.column for cell in header_row if str(cell.value).strip() in wanted]
    if not col_indexes:
        return False

    for col_idx in col_indexes:
        for (cell,) in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            cell.fill = HIGHLIGHT_FILL

    out_dir = os.path.join(_user_data_dir(), "highlighted_references")
    os.makedirs(out_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    safe_sheet = re.sub(r'[\\/*?:"<>|\[\]]', "_", sheet_name)[:30]
    out_path = os.path.join(out_dir, f"{base_name}__{safe_sheet}__highlighted.xlsx")

    try:
        wb.save(out_path)
    except Exception:
        return False

    return open_file(out_path)