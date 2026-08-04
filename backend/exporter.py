"""
exporter.py

Turns the app's in-memory data (consolidated rows, GST summary, party
summary, validation issues, filtered views) into a single, properly
formatted .xlsx workbook -- the kind a CA would actually want to open:
bold coloured headers, frozen header row, autofilter, sensible column
widths, number formatting on amount columns, and a bold Grand Total
row at the bottom of every data sheet.

Also builds a "Source Files" reference sheet (every sub file/company
that fed the report), a Credit = Debit checkpoint on the Summary
sheet, and turns each "Source File" cell on Consolidated Data into a
clickable link back to its row on the Source Files sheet.

Also provides open_file()/open_containing_folder() helpers so the UI
can take the user straight into Excel after export, cross-platform.
"""

import os
import re
import platform
import subprocess
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from backend.filters import numeric_columns, grand_totals
from backend.column_mapper import detect_columns
from backend.settings import _user_data_dir

# Text shown whenever a column has no real heading in the source file --
# kept as one constant so it's identical everywhere it's used.
NO_HEADING_TEXT = "No heading in source file"
LINK_FONT = Font(color="1155CC", underline="single", size=11)

HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16, color="1E3A8A")
HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"),
)

# One distinct colour per validation-issue type, so the "Issues Found"
# sheet is scannable at a glance -- e.g. every "Missing Required Fields"
# row is the same amber shade, every "Exact Duplicate Rows" row is the
# same blue shade, etc. (bg_hex, text_hex, label used in the legend).
ISSUE_TYPE_COLORS = {
    "missing_required_fields":    ("FEF3C7", "92400E", "Missing Required Fields"),
    "invalid_gstin":               ("FCE7F3", "9D174D", "Invalid GSTIN"),
    "exact_duplicate_rows":        ("DBEAFE", "1E40AF", "Exact Duplicate Rows"),
    "duplicate_invoice_numbers":   ("E0E7FF", "3730A3", "Duplicate Invoice Numbers"),
    "missing_invoice_numbers":     ("FEE2E2", "991B1B", "Missing Invoice Numbers"),
}
DEFAULT_ISSUE_COLOR = ("F3F4F6", "374151", "Other")

# Matches the "File: {name} | Sheet: {sheet} | {detail}" strings built
# in ui/dashboard.py._compute_all_issues, so the sheet name and file
# name can be pulled into their own columns instead of one long string.
_ISSUE_DETAIL_RE = re.compile(r"^File:\s*(.*?)\s*\|\s*Sheet:\s*(.*?)\s*\|\s*(.*)$")


def _to_number(value, default=0.0):
    try:
        if value == "" or value is None:
            return default
        return float(value)
    except (ValueError, TypeError):
        return default


def _credit_debit_checkpoint(rows, columns):
    """Checks whether Total Debit equals Total Credit across every
    consolidated row -- a common tie-out check before a report is
    considered final. Returns None if the source data doesn't have
    both a debit and a credit column, so the section is simply
    skipped rather than showing a false mismatch."""
    mapping = detect_columns(columns)
    debit_col = mapping.get("debit")
    credit_col = mapping.get("credit")
    if not debit_col or not credit_col:
        return None
    total_debit = sum(_to_number(row.get(debit_col)) for row in rows)
    total_credit = sum(_to_number(row.get(credit_col)) for row in rows)
    difference = round(total_debit - total_credit, 2)
    return {
        "debit_col": debit_col,
        "credit_col": credit_col,
        "total_debit": round(total_debit, 2),
        "total_credit": round(total_credit, 2),
        "difference": difference,
        "matched": abs(difference) < 0.01,
    }


_TITLE_DATE_RE = re.compile(
    r"(As of\s+[A-Za-z]+\s+\d{1,2},?\s+\d{4}|[A-Za-z]{3,9}\s+\d{1,2},?\s+\d{4}|\b[A-Za-z]{3}\s+\d{1,2},\s*\d{2}\b)"
)


def _title_rich_text(title):
    """Returns the report title/period text (e.g. 'GSR Foods II, LLC -
    Trial Balance - As of March 31, 2025 - Mar 31, 25') with any date
    portion rendered in bold, as an openpyxl rich-text value. Falls
    back to the plain string when there's no date to bold."""
    if not title:
        return title
    matches = list(_TITLE_DATE_RE.finditer(title))
    if not matches:
        return title
    bold_font = InlineFont(b=True)
    blocks = []
    pos = 0
    for m in matches:
        if m.start() > pos:
            blocks.append(title[pos:m.start()])
        blocks.append(TextBlock(bold_font, m.group()))
        pos = m.end()
    if pos < len(title):
        blocks.append(title[pos:])
    return CellRichText(*blocks)


def _source_files_breakdown(rows):
    """Groups consolidated rows by the file they came from, so each
    'sub file' that went into the report can be listed on its own
    reference sheet -- file name, company name assigned to it, how
    many rows it contributed, and the report title/period detected
    in the source file (e.g. "Trial Balance - As of March 31, 2025"),
    if any."""
    breakdown = {}
    order = []
    for row in rows:
        fname = row.get("__source_file", "Unknown")
        if fname not in breakdown:
            breakdown[fname] = {"company_name": row.get("Company Name", ""), "row_count": 0, "report_title": ""}
            order.append(fname)
        breakdown[fname]["row_count"] += 1
        if not breakdown[fname]["report_title"] and row.get("__report_title"):
            breakdown[fname]["report_title"] = row["__report_title"]
    return order, breakdown


def _autofit_columns(ws, columns, rows, min_width=10, max_width=45):
    for idx, col in enumerate(columns, start=1):
        longest = len(str(col))
        for row in rows[:500]:
            val_len = len(str(row.get(col, "")))
            if val_len > longest:
                longest = val_len
        ws.column_dimensions[get_column_letter(idx)].width = max(min_width, min(longest + 3, max_width))


GROUP_SHADE_FILLS = [
    PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
    PatternFill(start_color="EAF1FB", end_color="EAF1FB", fill_type="solid"),
]
BOUNDARY_TOP_SIDE = Side(style="medium", color="1E3A8A")


def _shade_source_file_groups(ws, rows, num_columns, start_row=1):
    """Purely cosmetic pass over the Consolidated Data sheet: doesn't
    touch any cell value, doesn't add/remove rows, and doesn't tint
    the background. It just draws a bold top border on the first row
    of every new source file/sheet block -- so it's visually obvious
    where one sheet's (or file's) data ends and the next one's begins,
    even when every row came from the same workbook (e.g. a Trial
    Balance sheet vs a Balance Sheet vs a P&L sheet all in one file)."""
    prev_group = None
    r = start_row
    for row in rows:
        r += 1
        group_key = (row.get("__source_file", ""), row.get("__source_sheet", ""))
        is_boundary = prev_group is not None and group_key != prev_group
        prev_group = group_key
        if is_boundary:
            for c_idx in range(1, num_columns + 1):
                cell = ws.cell(row=r, column=c_idx)
                existing = cell.border
                cell.border = Border(
                    left=existing.left, right=existing.right,
                    top=BOUNDARY_TOP_SIDE, bottom=existing.bottom,
                )


def _highlight_total_rows(ws, columns, rows, start_row=1):
    """A row whose Account cell literally reads "TOTAL" is a
    per-company subtotal carried over from the original source file
    (e.g. QuickBooks trial balances often have one). Left as-is it
    sits squeezed into one narrow column while the column right next
    to it is empty for that row. This merges the "TOTAL" label into
    that empty space beside it (so it reads across both columns
    instead of looking cramped) and draws a line directly above and
    below the row -- it does NOT touch the row's fill color, which
    stays whatever the source-file group shading already set it to."""
    r = start_row
    for row in rows:
        r += 1
        total_idx = None
        for idx, col in enumerate(columns):
            if str(row.get(col, "")).strip().upper() == "TOTAL":
                total_idx = idx
                break
        if total_idx is None:
            continue

        # Shift the label into the next column if it's empty there.
        if total_idx + 1 < len(columns):
            next_col = columns[total_idx + 1]
            if str(row.get(next_col, "")).strip() == "":
                first_c = total_idx + 1   # 1-based openpyxl column of the TOTAL cell
                second_c = total_idx + 2  # 1-based openpyxl column of the empty cell beside it
                ws.merge_cells(start_row=r, start_column=first_c, end_row=r, end_column=second_c)
                merged_cell = ws.cell(row=r, column=first_c)
                merged_cell.value = "TOTAL"
                merged_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Line above and below the row -- no fill/font change at all.
        for c_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c_idx)
            existing = cell.border
            cell.border = Border(
                left=existing.left, right=existing.right,
                top=BOUNDARY_TOP_SIDE, bottom=BOUNDARY_TOP_SIDE,
            )


_TOTAL_WORD_RE = re.compile(r"\btotal\b", re.IGNORECASE)


def _bold_subtotal_label_rows(ws, columns, rows, start_row=1):
    """Some source files carry subtotal lines that are worded like
    "Total Cash - Checking", "Total Checking/Savings", "Total Current
    Assets", etc. -- the word "TOTAL" is only part of the cell's text,
    not the whole cell, so `_highlight_total_rows` (exact-match only)
    skips them and they land in the export looking like an ordinary
    line item even though they're really a rolled-up figure carried
    over from a different part of the source sheet.

    This bolds the label cell (so "Total" stands out) and bolds
    whichever cell(s) in that same row hold the actual number, so the
    user can immediately see it's a subtotal, not a regular row --
    without touching fill color, borders, or any other row already
    handled by `_highlight_total_rows`."""
    r = start_row
    for row in rows:
        r += 1
        label_idx = None
        for idx, col in enumerate(columns):
            text = str(row.get(col, "")).strip()
            if text.upper() == "TOTAL":
                # Exact match -- already handled elsewhere.
                label_idx = None
                break
            if _TOTAL_WORD_RE.search(text):
                label_idx = idx
                break
        if label_idx is None:
            continue

        for c_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=c_idx)
            value = row.get(col, "")
            is_label = (c_idx - 1) == label_idx
            is_amount = value not in ("", None) and not is_label
            if is_label or is_amount:
                existing = cell.font
                cell.font = Font(
                    name=existing.name, size=existing.size, bold=True,
                    italic=existing.italic, color=existing.color,
                )


def _bold_company_name_column(ws, columns, num_rows, start_row=1):
    """Bolds every cell in the "Company Name" column on the
    Consolidated Data sheet (header already bold from the header
    style -- this covers the data rows below it), so the company each
    row belongs to stands out at a glance. Purely cosmetic -- doesn't
    touch any other column, fill, or border."""
    if "Company Name" not in columns:
        return
    col_idx = columns.index("Company Name") + 1
    for r in range(start_row + 1, start_row + 1 + num_rows):
        cell = ws.cell(row=r, column=col_idx)
        existing = cell.font
        cell.font = Font(
            name=existing.name, size=existing.size, bold=True,
            italic=existing.italic, color=existing.color,
        )


def _write_table(ws, columns, rows, start_row=1, add_grand_total=True):
    for c_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=start_row, column=c_idx, value=str(col))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    num_cols = numeric_columns(rows, columns, sample_size=len(rows) or 1)

    r = start_row
    for row in rows:
        r += 1
        for c_idx, col in enumerate(columns, start=1):
            value = row.get(col, "")
            cell = ws.cell(row=r, column=c_idx, value=value)
            cell.border = THIN_BORDER
            if col in num_cols and value not in ("", None):
                try:
                    cell.value = float(value)
                    cell.number_format = "#,##,##0.00"
                except (ValueError, TypeError):
                    pass

    if add_grand_total and rows:
        r += 1
        totals = grand_totals(rows, columns)
        label_written = False
        for c_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=c_idx)
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            cell.border = THIN_BORDER
            if col in totals:
                cell.value = totals[col]
                cell.number_format = "#,##,##0.00"
            elif not label_written:
                cell.value = "GRAND TOTAL"
                label_written = True

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    if rows:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(columns))}{start_row + len(rows)}"
    _autofit_columns(ws, columns, rows)
    return r + 2


def export_full_report(output_path, consolidated, gst_summary=None, party_summary=None,
                        validation_issues=None, source_files=None, company_summary=None,
                        month_summary=None, consolidated_at=None, sections=None, file_paths=None):
    """
    Builds the complete multi-sheet workbook:
      1. Index               - what's in this workbook and why (so nobody
                                has to guess what each tab is for)
      2. Summary              - headline numbers, which files went in, and
                                the Credit = Debit checkpoint
      3. Source Files         - every sub file merged in, with its Company
                                Name and row count (linked from Consolidated Data)
      4. Consolidated Data    - every merged row + Grand Total row, with a
                                clickable Source File reference
      5. Company Summary      - totals per company (from the always-filled
                                 "Company Name" column)
      6. Issues Found         - every validation problem, with reference
    Any section is skipped gracefully if its data wasn't supplied.
    """

    sections = sections or {}
    file_paths = file_paths or {}

    wb = Workbook()
    rows = consolidated.get("rows", [])
    columns = [c for c in consolidated.get("columns", []) if not c.startswith("__")]
    # Guard against any stray blank header slipping through to export --
    # every column must show something a user can understand.
    columns = [c if str(c).strip() else NO_HEADING_TEXT for c in columns]
    # Safety net: a column with no real heading is never shown in the
    # export, full stop -- even if it somehow reached this point still
    # carrying data (e.g. from an already-imported session predating an
    # import-side fix). Drop the column and strip it out of every row.
    no_heading_cols = [c for c in columns if "no heading" in str(c).lower()]
    if no_heading_cols:
        columns = [c for c in columns if c not in no_heading_cols]
        rows = [{k: v for k, v in row.items() if k not in no_heading_cols} for row in rows]
    # Safety net #2: a column can look fine in any one file on its own
    # (so the per-file import cleanup leaves it alone) but still turn
    # out completely blank once every file is merged together -- e.g.
    # a column named "Account" that no source file in this batch ever
    # actually put data into. Check across the WHOLE merged dataset and
    # drop any column that's empty in every single row, from every file.
    if rows:
        fully_empty_cols = [
            c for c in columns
            if all(str(row.get(c, "")).strip() == "" for row in rows)
        ]
        if fully_empty_cols:
            columns = [c for c in columns if c not in fully_empty_cols]
            rows = [{k: v for k, v in row.items() if k not in fully_empty_cols} for row in rows]
    # Company Name always leads the sheet since every row is tagged with it.
    if "Company Name" in columns:
        columns = ["Company Name"] + [c for c in columns if c != "Company Name"]
    # "Account" is redundant/mostly blank whenever a more specific
    # "Account (2)" column is also present -- drop it so the TOTAL row
    # label doesn't visually land in the wrong column.
    if "Account" in columns and "Account (2)" in columns:
        # Merge per row instead of dropping one column outright -- some
        # source files only put their real account label in "Account"
        # (no "Account (2)" at all in that file), others only put it in
        # "Account (2)". Keep whichever one actually has a value.
        for row in rows:
            merged_val = row.get("Account (2)") or row.get("Account") or ""
            row["Account"] = merged_val
            row.pop("Account (2)", None)
        columns = [c for c in columns if c != "Account (2)"]
    has_sheet = bool(rows) and any(row.get("__source_sheet") for row in rows)
    display_columns = (
        (["Source Sheet"] if has_sheet else [])
        + columns
    )

    # Index / Summary / Source Files sheets removed per request --
    # only Consolidated Data (+ Issues Found, if any) are exported.
    # file_row_map is kept as an empty dict so the Issues Found section
    # below (unchanged) still runs safely without a Source Files sheet
    # to link to -- it just won't have a hyperlink target.
    file_row_map = {}
    wb.remove(wb.active)

    # ---------- 4. Consolidated Data sheet ----------
    ws2 = wb.create_sheet("Consolidated Data")
    export_rows = []
    for row in rows:
        flat = {c: row.get(c, "") for c in columns}
        if "Source File" in display_columns:
            flat["Source File"] = row.get("__source_file", "")
        if "Source Sheet" in display_columns:
            flat["Source Sheet"] = row.get("__source_sheet", "")
        export_rows.append(flat)
    last_data_row = _write_table(ws2, display_columns, export_rows, add_grand_total=False)

    # Cosmetic only -- mark where a new file/sheet's data starts, so
    # different files are easy to tell apart.
    _shade_source_file_groups(ws2, rows, len(display_columns))

    # Cosmetic only -- bold the Company Name column so it stands out.
    _bold_company_name_column(ws2, display_columns, len(export_rows))

    # Cosmetic only -- per-company "TOTAL" rows get their label merged
    # into the empty cell beside it and the whole row highlighted.
    _highlight_total_rows(ws2, display_columns, rows)

    # Cosmetic only -- rows like "Total Cash - Checking" or "Total
    # Current Assets" carry the word "Total" inside a longer label
    # rather than as the whole cell, so the row above misses them.
    # Bold the label and its amount so these subtotal rows are easy
    # to spot instead of blending in with ordinary line items.
    _bold_subtotal_label_rows(ws2, display_columns, rows)

    # Turn each "Source File" cell into a clickable reference back to
    # that file's row on the Source Files sheet.
    if "Source File" in display_columns and file_row_map:
        src_col_idx = display_columns.index("Source File") + 1
        for r_idx, row in enumerate(export_rows, start=2):
            fname = row.get("Source File", "")
            target_row = file_row_map.get(fname)
            if target_row:
                cell = ws2.cell(row=r_idx, column=src_col_idx)
                cell.hyperlink = f"#'Source Files'!B{target_row}"
                cell.font = LINK_FONT

    # ---------- 6. Issues Found sheet ----------
    if validation_issues:
        ws5 = wb.create_sheet("Issues Found")

        # ---- Legend: what each colour means, so it's self-explanatory ----
        ws5.cell(row=1, column=1, value="Colour Guide:").font = Font(bold=True, size=11)
        used_types = [t for t in validation_issues.keys() if validation_issues.get(t)]
        legend_col = 2
        for issue_type in used_types:
            bg, fg, label = ISSUE_TYPE_COLORS.get(issue_type, DEFAULT_ISSUE_COLOR)
            c = ws5.cell(row=1, column=legend_col, value=f"  {label}  ")
            c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            c.font = Font(bold=True, color=fg, size=10)
            c.alignment = Alignment(horizontal="center")
            legend_col += 1

        header_row = 3
        columns5 = ["Issue Type", "Source File", "Sheet", "Details"]
        for c_idx, col in enumerate(columns5, start=1):
            cell = ws5.cell(row=header_row, column=c_idx, value=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        r5 = header_row
        any_rows = False
        for issue_type, items in validation_issues.items():
            if not items:
                continue
            bg, fg, label = ISSUE_TYPE_COLORS.get(issue_type, DEFAULT_ISSUE_COLOR)
            row_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            row_font = Font(color=fg, size=11)
            for item in items:
                any_rows = True
                r5 += 1
                text = str(item)
                m = _ISSUE_DETAIL_RE.match(text)
                if m:
                    fname, sheet_name, detail = m.group(1), m.group(2), m.group(3)
                else:
                    fname, sheet_name, detail = "", "", text

                type_cell = ws5.cell(row=r5, column=1, value=label)
                src_cell = ws5.cell(row=r5, column=2, value=fname)
                sheet_cell = ws5.cell(row=r5, column=3, value=sheet_name)
                detail_cell = ws5.cell(row=r5, column=4, value=detail)

                for cell in (type_cell, src_cell, sheet_cell, detail_cell):
                    cell.fill = row_fill
                    cell.font = row_font
                    cell.border = THIN_BORDER

                # Clicking the source file jumps to its row on "Source
                # Files" (if that sheet exists) or opens it on disk.
                target_row = file_row_map.get(fname)
                src_path = file_paths.get(fname)
                if target_row:
                    src_cell.hyperlink = f"#'Source Files'!B{target_row}"
                    src_cell.font = Font(color="1155CC", underline="single", size=11)
                elif src_path:
                    src_cell.hyperlink = f"file:///{src_path}"
                    src_cell.font = Font(color="1155CC", underline="single", size=11)

        if any_rows:
            ws5.freeze_panes = ws5.cell(row=header_row + 1, column=1)
            ws5.auto_filter.ref = f"A{header_row}:D{r5}"
            ws5.column_dimensions["A"].width = 26
            ws5.column_dimensions["B"].width = 32
            ws5.column_dimensions["C"].width = 22
            ws5.column_dimensions["D"].width = 55
        else:
            ws5.cell(row=header_row + 1, column=1, value="No issues found").font = Font(bold=True, color="16A34A")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


def export_filtered_view(output_path, columns, rows, title="Filtered Data"):
    """Simpler single-sheet export used by the Filters screen and the
    Analytics screen (exporting whatever the user is currently looking at)."""
    wb = Workbook()
    ws = wb.active
    ws.title = (title[:31] if title else "Filtered Data")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated on: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}  |  {len(rows)} row(s)"
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    display_columns = [c if str(c).strip() else NO_HEADING_TEXT for c in columns if not str(c).startswith("__")]
    _write_table(ws, display_columns, rows, start_row=4)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


def export_pivot_view(output_path, pivot_columns, pivot_rows, title="Pivot Table"):
    """Exports a pivot-shaped table (list of flat dict rows) with the
    same styling as everything else."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pivot Table"
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated on: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    _write_table(ws, pivot_columns, pivot_rows, start_row=4, add_grand_total=False)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


# ---------------- Cross-platform "open file / folder" helpers ---------------- #

def open_file(path):
    """Opens the given file with whatever application the OS has
    associated with .xlsx (i.e. Excel, if installed) -- this is what
    takes the user directly into Excel after export."""
    try:
        system = platform.system()
        if system == "Windows":
            os.startfile(path)  # noqa: only exists on Windows
        elif system == "Darwin":
            subprocess.run(["open", path], check=False)
        else:
            subprocess.run(["xdg-open", path], check=False)
        return True
    except Exception:
        return False


def open_containing_folder(path):
    """Opens the folder containing the file (selecting the file itself
    where the OS supports it)."""
    folder = os.path.dirname(os.path.abspath(path))
    try:
        system = platform.system()
        if system == "Windows":
            subprocess.run(["explorer", "/select,", os.path.abspath(path)], check=False)
        elif system == "Darwin":
            subprocess.run(["open", "-R", os.path.abspath(path)], check=False)
        else:
            subprocess.run(["xdg-open", folder], check=False)
        return True
    except Exception:
        return False


def highlighted_copy_path(file_path, sheet_name):
    """Returns the deterministic path of the highlighted reference copy
    that open_file_with_highlighted_columns() creates for a given
    source file + sheet (same folder, same naming pattern it always
    uses). Doesn't check whether the file actually exists yet -- callers
    that need that should os.path.exists() the result themselves.

    Exposed as its own function (rather than only inline inside
    open_file_with_highlighted_columns) so the Debit/Credit Checkpoint's
    'Refresh' button can look for a saved, corrected copy in the same
    place 'Open File' would have written one, without duplicating the
    naming logic."""
    out_dir = os.path.join(_user_data_dir(), "highlighted_references")
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    safe_sheet = re.sub(r'[\\/*?:"<>|\[\]]', "_", sheet_name)[:30]
    return os.path.join(out_dir, f"{base_name}__{safe_sheet}__highlighted.xlsx")


def open_file_with_highlighted_columns(file_path, sheet_name, columns_to_highlight):
    """Used by the Debit/Credit Checkpoint screen: makes a copy of the
    given source file with the named column(s) highlighted in yellow
    on the given sheet, then opens that copy in Excel -- so a mismatch
    can be traced straight back to the exact columns it came from.

    A copy is opened rather than the original file so the user's
    source file on disk is never modified. Copies are written to
    <app data>/highlighted_references and reused/overwritten on
    repeat opens of the same file/sheet/columns combination.

    Returns a (success, reason) tuple. success is True if the
    highlighted copy was created and opened; reason is None on
    success, or a short human-readable message on failure explaining
    why the source file, sheet, or columns could not be found.
    """
    if not file_path or not os.path.exists(file_path):
        return False, "The source file could not be found on disk."

    wanted = {str(c).strip() for c in (columns_to_highlight or []) if c}
    if not wanted:
        return False, "No Debit/Credit columns were specified to highlight."

    try:
        wb = load_workbook(file_path)
    except Exception as e:
        return False, f"Couldn't open the workbook: {e}"

    if sheet_name not in wb.sheetnames:
        return False, f"Sheet '{sheet_name}' was not found in this file."
    ws = wb[sheet_name]

    header_row = next(ws.iter_rows(min_row=1, max_row=1), ())
    col_indexes = [cell.column for cell in header_row if str(cell.value).strip() in wanted]
    if not col_indexes:
        return False, "None of the Debit/Credit columns were found in the sheet header."

    for col_idx in col_indexes:
        for (cell,) in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            cell.fill = HIGHLIGHT_FILL

    out_dir = os.path.join(_user_data_dir(), "highlighted_references")
    os.makedirs(out_dir, exist_ok=True)
    out_path = highlighted_copy_path(file_path, sheet_name)

    try:
        wb.save(out_path)
    except PermissionError:
        # The existing copy is most likely still open in Excel (Windows
        # locks open files), so overwriting it fails with Errno 13.
        # Fall back to a fresh, uniquely-named copy instead of failing
        # outright -- the user doesn't need to close their existing
        # window just to look at the reference again.
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        safe_sheet = re.sub(r'[\\/*?:"<>|\[\]]', "_", sheet_name)[:30]
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(out_dir, f"{base_name}__{safe_sheet}__highlighted_{stamp}.xlsx")
        try:
            wb.save(out_path)
        except Exception as e:
            return False, (
                "Couldn't save the highlighted copy because the file appears to be "
                f"open elsewhere (e.g. in Excel). Close it and try again. ({e})"
            )
    except Exception as e:
        return False, f"Couldn't save the highlighted copy: {e}"

    if open_file(out_path):
        return True, None
    return False, "The highlighted copy was saved but couldn't be opened automatically."